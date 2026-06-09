from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

from flask import current_app
from openpyxl import load_workbook

from services.text_utils import normalizar_texto, valor_verdadeiro


class FamiliasRepository:
    def buscar_familia(self, codigo_marca: str, codigo_modelo: str, marca: str, modelo: str) -> dict[str, Any] | None:
        registros = self.carregar_familias()
        marca_id = str(codigo_marca or "").strip()
        modelo_id = str(codigo_modelo or "").strip()

        if marca_id and modelo_id:
            for item in registros:
                if str(item.get("marca_id", "")).strip() == marca_id and str(item.get("modelo_id", "")).strip() == modelo_id:
                    return item

        marca_norm = normalizar_texto(marca)
        modelo_norm = normalizar_texto(modelo)
        if modelo_norm:
            for item in registros:
                if marca_norm and normalizar_texto(item.get("marca_nome", "")) != marca_norm:
                    continue
                modelo_item = normalizar_texto(item.get("modelo_nome", ""))
                if modelo_norm == modelo_item or modelo_norm in modelo_item or modelo_item in modelo_norm:
                    return item
        return None

    def carregar_familias(self) -> list[dict[str, Any]]:
        caminho = Path(current_app.config["ARQUIVO_FAMILIAS"])
        return list(self._carregar_familias_cache(str(caminho)))

    @staticmethod
    @lru_cache(maxsize=4)
    def _carregar_familias_cache(caminho_str: str) -> tuple[dict[str, Any], ...]:
        caminho = Path(caminho_str)
        if not caminho.exists():
            return tuple()

        wb = load_workbook(caminho, data_only=True, read_only=True)
        registros: list[dict[str, Any]] = []

        for ws in wb.worksheets:
            if str(ws.title).strip().upper() == "CONTROLE":
                continue

            linhas = ws.iter_rows(values_only=True)
            headers = next(linhas, None)
            if not headers:
                continue
            idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}

            if "marca_id" not in idx or "modelo_id" not in idx:
                continue

            for row in linhas:
                def v(nome: str) -> str:
                    pos = idx.get(nome)
                    if pos is None or pos >= len(row):
                        return ""
                    valor = row[pos]
                    return "" if valor is None else str(valor).strip()

                marca_id = v("marca_id")
                modelo_id = v("modelo_id")
                modelo_nome = v("modelo_nome")
                if not marca_id or not modelo_id or not modelo_nome:
                    continue

                registros.append({
                    "family_id": v("family_id"),
                    "family_nome": v("family_nome") or v("family_name"),
                    "marca_id": marca_id,
                    "marca_nome": v("marca_nome") or ws.title,
                    "modelo_id": modelo_id,
                    "modelo_nome": modelo_nome,
                    "subtipo": v("subtipo"),
                    "ativo": v("ativo"),
                    "curva_pronta": v("curva_pronta"),
                    "modelo_base_curva": v("modelo_base_curva"),
                    "ano_base_curva": v("ano_base_curva"),
                    "curva_pronta_combustao": v("curva_pronta_combustao"),
                    "historico_baixado_combustao": v("historico_baixado_combustao"),
                    "modelo_base_curva_combustao": v("modelo_base_curva_combustao"),
                    "ano_base_curva_combustao": v("ano_base_curva_combustao"),
                    "curva_pronta_eletrico": v("curva_pronta_eletrico"),
                    "historico_baixado_eletrico": v("historico_baixado_eletrico"),
                    "modelo_base_curva_eletrico": v("modelo_base_curva_eletrico"),
                    "ano_base_curva_eletrico": v("ano_base_curva_eletrico"),
                })

        return tuple(registros)
